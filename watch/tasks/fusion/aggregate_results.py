"""
Loads and summarizes pre-computed metrics over multiple experiments

TODO:

    AUTO-ABALATION

    - [ ] Rename to aggregate results

    - [ ] Group experiments into "variance" groups where params are exactly the
          same in all members of the group.

    - [ ] Group experiments into 1-abalateable groups where params are only 1
          param is different between any two members in the group.

    - [ ] Group experiments into 2-abalateable groups where params are only 2
          params are different between any two members in the group. (must be the
          same param between all members of the groups).


The main function is :func:`main`.

CommandLine:

    # TODO: - [ ] watch command to "register" which DVC directory to use
    # Locate a registered DVC directory
    DVC_DPATH=$(smartwatch_dvc)

    # Replace with the name of the latest dataset
    DATASET_CODE=Drop2-Aligned-TA1-2022-02-15
    EXPT_NAME_PAT="*"
    #EXPT_NAME_PAT="BOTH_TA1_COMBO_TINY_p2w_raw*"
    MODEL_EPOCH_PAT="*"
    PRED_DSET_PAT="*"
    MEASURE_GLOBSTR=${DVC_DPATH}/models/fusion/${DATASET_CODE}/${EXPT_NAME_PAT}/${MODEL_EPOCH_PAT}/${PRED_DSET_PAT}/eval/curves/measures2.json

    python -m watch.tasks.fusion.aggregate_results \
        --measure_globstr="$MEASURE_GLOBSTR" \
        --out_dpath="$DVC_DPATH/agg_results/" \
        --dset_group_key="*_vali.kwcoco" --show=True
"""
import json
import pandas as pd
import numpy as np
import ubelt as ub
import yaml
import io
import shutil
import kwarray
import warnings
import scriptconfig as scfg
import itertools as it


class AggregateResultsConfig(scfg.Config):
    """
    TODO: write good docs for the gather command line tool.

    Basic idea:

        Grabs a selection of precomputed metrics on a particular dataset

        Compares models against each other with statistical tests and plots.

        Tries to figure out which configs did best.
    """
    default = {
        'measure_globstr': scfg.Value('measures2.json', help='a group of measures2.json files from kwcoco metrics, specified by list or glob pattern'),
        'out_dpath': scfg.Value('./agg_results', help='A location where aggregate results can be written and compared'),
        'show': scfg.Value(False, help='if true, does a plt.show'),
        'dset_group_key': scfg.Value('*', help='if there is more than one dataset group, you will need to choose one'),

        'io_workers': scfg.Value(10, help='number of workers to read metrics summaries'),

        'classes_of_interest': scfg.Value('*', nargs='+', help='One or more glob patterns'),

        'embed': scfg.Value(False, help='if true embed into IPython before viz'),

        'max_per_expt_salient_curves': 3,
    }


class Found(Exception):
    pass


def _writefig(fig, dpath, fname, figsize, verbose, tight):
    fig_fpath = dpath / fname
    if verbose:
        print('write fig_fpath = {!r}'.format(fig_fpath))
    fig.set_size_inches(figsize)
    if tight:
        fig.tight_layout()
    fig.savefig(fig_fpath)


def debug_all_results():
    """
    ls models/fusion/eval3_candidates/eval/*/*/*/*/eval/tracking/*/iarpa_eval/scores/merged/summary2.json
    ls models/fusion/eval3_sc_candidates/eval/*/*/*/*/eval/actclf/*/*_eval/scores/merged/summary3.json

    ls models/fusion/eval3_sc_candidates/pred/*/*/*/*/actclf/*/*_eval/scores/merged/summary3.json

    models/fusion/eval3_candidates/eval/*/*/*/*/eval/tracking/

    find models/fusion -iname "summary2.json"
    find models/fusion -iname "summary3.json"
    """

    import watch
    from watch.utils import util_path
    import pandas as pd
    dvc_dpath = watch.find_smart_dvc_dpath(hardware='hdd')
    ignore_cols = [
        'modulate_class_weights', 'accumulate_grad_batches',
        'pred_in_dataset_fpath', 'pred_model_fpath'
        'name',
        'patience', 'normalize_inputs',
        'neg_to_pos_ratio',
        'use_special_classes', 'pred_model_fpath',
        'temporal_dropout',
        'use_centered_positives',
        'chip_overlap',
        'arch_name',
        'class_loss', 'global_change_weight',
        'learning_rate',
        'max_epoch_length',
        'stream_channels',
        'name',
        'time_sampling',
        'max_epochs',
        'dist_weights',
        'saliency_loss',
        'global_class_weight',
        'init',
    ]

    bas_globpats = [
        dvc_dpath / 'models/fusion/eval3_candidates/eval/*/*/*/*/eval/tracking/*/iarpa_eval/scores/merged/summary2.json'
    ]
    bas_paths = util_path.coerce_patterned_paths(bas_globpats)
    from watch.utils import result_analysis

    bas_rows = []
    for merged_fpath in bas_paths:
        with open(merged_fpath, 'r') as file:
            bas_info = json.load(file)

        best_bas_rows = pd.read_json(io.StringIO(json.dumps(bas_info['best_bas_rows'])), orient='table')
        try:
            bas_row = best_bas_rows.loc['merged'].reset_index().iloc[0].to_dict()
        except Exception:
            bas_row = best_bas_rows[best_bas_rows['region_id'].isnull()].reset_index(drop=1).iloc[0].to_dict()

        tracker_info = bas_info['parent_info']
        path_hint = merged_fpath
        params = parse_tracker_params(tracker_info, dvc_dpath, path_hint=path_hint)

        metrics = {
            'BAS_F1': bas_row['F1'],
            'BAS_rho': bas_row['rho'],
            'BAS_tau': bas_row['tau'],
            # 'mean_f1': sc_df.loc['F1 score'].mean(),
            # 'siteprep_f1': sc_df.loc['F1 score', 'Site Preparation'].mean(),
            # 'active_f1': sc_df.loc['F1 score', 'Active Construction'].mean(),
        }
        row = ub.odict(ub.dict_union(metrics, params))
        bas_rows.append(row)

    print(f'{len(bas_rows)=}')
    bas_rows = list(ub.unique(bas_rows, key=ub.hash_data))
    print(f'{len(bas_rows)=}')

    df = pd.DataFrame(bas_rows)
    df = df.sort_values('BAS_F1')

    varied = ub.varied_values(bas_rows, 1, None)
    varied2 = {k: v for k, v in varied.items() if len(ub.oset(v) - {None}) > 1}
    varied_cols = ub.oset(df.columns) & list(varied2.keys())
    df2 = df[varied_cols].sort_values('BAS_F1')
    df2 = shrink_notations(df2)
    df2 = df2.drop(ub.oset(df2.columns) & ignore_cols, axis=1)
    print(df2.to_string())
    # print(df2.iloc[-70:].to_string())

    ###
    ###
    ###

    sc_globpats = [
        # dvc_dpath / 'models/fusion/eval3_sc_candidates/pred/*/*/*/*/actclf/*/*_eval/scores/merged/summary3.json',
        # dvc_dpath / 'models/fusion/eval3_sc_candidates/eval/*/*/*/*/eval/actclf/*/*_eval/scores/merged/summary3.json'

        dvc_dpath / 'models/fusion/eval3_sc_candidates/eval/CropDrop3_SC_V006/pred_CropDrop3_SC_V006_epoch=71-step=18431/*/*/eval/actclf/*/*_eval/scores/merged/summary3.json'

        # dvc_dpath / 'models/fusion/eval3_sc_candidates/pred/*/*/*/*/actclf/*/iarpa_sc_eval/scores/merged/summary3.json',
    ]
    sc_paths = util_path.coerce_patterned_paths(sc_globpats)

    seen = set()
    sc_rows = []
    sc_cms = []
    result_list = []
    for merged_fpath in sc_paths:
        with open(merged_fpath, 'r') as file:
            sc_info = json.load(file)
        # sc_info['sc_cm']
        sc_df = pd.read_json(io.StringIO(json.dumps(sc_info['sc_df'])), orient='table')
        sc_cm = pd.read_json(io.StringIO(json.dumps(sc_info['sc_cm'])), orient='table')
        sc_cms.append(sc_cm)
        tracker_info = sc_info['parent_info']
        param_types = parse_tracker_params(tracker_info, dvc_dpath)

        non_measures = ub.dict_diff(param_types, ['resource'])

        params = ub.dict_union(*non_measures.values())
        metrics = {
            'mean_f1': sc_df.loc['F1 score'].mean(),
            'siteprep_f1': sc_df.loc['F1 score', 'Site Preparation'].mean(),
            'active_f1': sc_df.loc['F1 score', 'Active Construction'].mean(),
        }
        metrics.update(
            param_types['resource']
        )
        row = ub.odict(ub.dict_union(metrics, *param_types.values()))

        actcfg_dname = merged_fpath.parent.parent.parent.parent.name
        predcfg_dname = merged_fpath.parent.parent.parent.parent.parent.parent.parent.name

        result = result_analysis.Result(
             name=None,
             params=params,
             metrics=metrics,
             meta=None
        )
        row['actcfg_dname'] = actcfg_dname
        row['predcfg_dname'] = predcfg_dname

        key = ub.hash_data(row)
        if key not in seen:
            sc_rows.append(row)
            sc_cms.append(sc_cm)
            result_list.append(result)
        seen.add(key)

    if 0:
        ub.util_hash._HASHABLE_EXTENSIONS.register(ub.Path)(lambda x: (b'path', ub.hash_data(str(x)).encode()))

    analysis = result_analysis.ResultAnalysis(
        result_list,
        # ignore_params=ignore_params,
        # metrics=['coi_mAPUC', 'coi_APUC'],
        # metrics=['salient_AP'],
        metrics=['mean_f1'],
        metric_objectives={
            'mean_f1': 'max',
        },
        # ignore_metrics=ignore_metrics,
        abalation_orders={1}
    )
    analysis.analysis()

    print(f'{len(sc_rows)=}')

    df = pd.DataFrame(sc_rows)
    df = df.sort_values('mean_f1')

    df[['trk_use_viterbi']]

    varied = ub.varied_values(sc_rows, 1, None)
    varied2 = {k: v for k, v in varied.items() if len(ub.oset(v) - {None}) > 1}
    varied_cols = ub.oset(df.columns) & list(varied2.keys())
    df2 = df[varied_cols].sort_values('mean_f1')
    df2 = shrink_notations(df2)
    df2 = df2.drop(ub.oset(df2.columns) & ignore_cols, axis=1)
    print(df2.reset_index().iloc[-90:].to_string())
    print(df2.reset_index().to_string())


@ub.memoize
def _load_json(fpath):
    # memo hack for development
    with open(fpath, 'r') as file:
        data = json.load(file)
    return data


def load_pxl_eval(fpath, dvc_dpath=None):
    from kwcoco.coco_evaluator import CocoSingleResult
    from watch.utils import util_pattern
    # from watch.utils import result_analysis
    # from watch.utils import util_time
    measure_info = _load_json(fpath)

    meta = measure_info['meta']

    pred_info = meta['info']
    param_types = parse_pred_params(pred_info, dvc_dpath)

    predict_args = param_types['pred']
    if predict_args is None:
        raise Exception('no prediction metadata')

    salient_measures = measure_info['nocls_measures']
    class_measures = measure_info['ovr_measures']

    # HACK: fixme
    coi_pattern = util_pattern.MultiPattern.coerce(
        ['Active Construction', 'Site Preparation'], hint='glob')

    class_metrics = []
    coi_metrics = []
    for catname, bin_measure in class_measures.items():
        class_row = {}
        class_row['AP'] = bin_measure['ap']
        class_row['AUC'] = bin_measure['auc']
        class_row['APUC'] = np.nanmean([bin_measure['ap'], bin_measure['auc']])
        class_row['catname'] = catname
        if coi_pattern.match(catname):
            coi_metrics.append(class_row)
        class_metrics.append(class_row)

    class_aps = [r['AP'] for r in class_metrics]
    class_aucs = [r['AUC'] for r in class_metrics]
    coi_aps = [r['AP'] for r in coi_metrics]
    coi_aucs = [r['AUC'] for r in coi_metrics]
    coi_catnames = [r['catname'] for r in coi_metrics]

    metrics = {}
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', 'Mean of empty slice')
        metrics['class_mAP'] = np.nanmean(class_aps) if len(class_aps) else np.nan
        metrics['class_mAUC'] = np.nanmean(class_aucs) if len(class_aucs) else np.nan
        metrics['class_mAPUC'] = np.nanmean([metrics['class_mAUC'], metrics['class_mAP']])

        metrics['coi_mAP'] = np.nanmean(coi_aps) if len(coi_aps) else np.nan
        metrics['coi_mAUC'] = np.nanmean(coi_aucs) if len(coi_aucs) else np.nan
        metrics['coi_mAPUC'] = np.nanmean([metrics['coi_mAUC'], metrics['coi_mAP']])

        metrics['salient_AP'] = salient_measures['ap']
        metrics['salient_AUC'] = salient_measures['auc']
        metrics['salient_APUC'] = np.nanmean([metrics['salient_AP'], metrics['salient_AUC']])

    for class_row in class_metrics:
        metrics[class_row['catname'] + '_AP'] = class_row['AP']
        metrics[class_row['catname'] + '_AUC'] = class_row['AUC']

    result = CocoSingleResult.from_json(measure_info)

    info = {
        'fpath': fpath,
        'metrics': metrics,
        'param_types': param_types,
        'other': {
            'result': result,
            'coi_catnames': ','.join(sorted(coi_catnames)),
            # 'sc_cm': sc_cm,
            # 'sc_df': sc_df,
        },
        'json_info': measure_info,
    }
    return info


def load_bas_eval(fpath, expt_dvc_dpath):
    bas_info = _load_json(fpath)

    best_bas_rows = pd.read_json(io.StringIO(json.dumps(bas_info['best_bas_rows'])), orient='table')

    flags = best_bas_rows['region_id'] == '__merged__'

    if np.any(flags):
        bas_row = best_bas_rows[flags].iloc[0]
    else:
        # OLD Phase 1 code, can eventually remove
        try:
            bas_row = best_bas_rows.loc['merged'].reset_index().iloc[0].to_dict()
        except Exception:
            bas_row = best_bas_rows[best_bas_rows['region_id'].isnull()].reset_index(drop=1).iloc[0].to_dict()

    tracker_info = bas_info['parent_info']
    path_hint = fpath
    param_types = parse_tracker_params(tracker_info, expt_dvc_dpath, path_hint=path_hint)

    metrics = {
        'BAS_F1': bas_row['F1'],
        'BAS_rho': bas_row['rho'],
        'BAS_tau': bas_row['tau'],
        # 'mean_f1': sc_df.loc['F1 score'].mean(),
        # 'siteprep_f1': sc_df.loc['F1 score', 'Site Preparation'].mean(),
        # 'active_f1': sc_df.loc['F1 score', 'Active Construction'].mean(),
    }
    info = {
        'fpath': fpath,
        'metrics': metrics,
        'param_types': param_types,
        'other': {
            'best_bas_rows': best_bas_rows,
        },
        'json_info': bas_info,
    }
    return info


def load_sc_eval(fpath, expt_dvc_dpath):
    sc_info = _load_json(fpath)
    # sc_info['sc_cm']
    sc_df = pd.read_json(io.StringIO(json.dumps(sc_info['sc_df'])), orient='table')
    sc_cm = pd.read_json(io.StringIO(json.dumps(sc_info['sc_cm'])), orient='table')
    tracker_info = sc_info['parent_info']
    param_types = parse_tracker_params(tracker_info, expt_dvc_dpath)

    # non_measures = ub.dict_diff(param_types, ['resource'])
    # params = ub.dict_union(*non_measures.values())
    metrics = {
        'mean_f1': sc_df.loc['F1 score'].mean(),
        'siteprep_f1': sc_df.loc['F1 score', 'Site Preparation'].mean(),
        'active_f1': sc_df.loc['F1 score', 'Active Construction'].mean(),
    }
    # metrics.update(
    #     param_types['resource']
    # )
    # row = ub.odict(ub.dict_union(metrics, *param_types.values()))

    info = {
        'fpath': fpath,
        'metrics': metrics,
        'param_types': param_types,
        'other': {
            'sc_cm': sc_cm,
            'sc_df': sc_df,
        },
        'json_info': sc_info,
    }
    return info


def parse_tracker_params(tracker_info, expt_dvc_dpath, path_hint=None):
    track_item = find_track_item(tracker_info)

    if 'extra' in track_item['properties']:
        pred_info = track_item['properties']['extra']['pred_info']
    elif 'pred_info' in track_item['properties']:
        pred_info = track_item['properties']['pred_info']
    else:
        if path_hint is None:
            raise Exception('cannot find pred info. This is an old result')
        # TODO: remove the eval stealing
        eval_dpath = path_hint.parent.parent.parent.parent.parent.parent
        # Can we steal pred info from pixel metrics?
        # eval_dpath / 'curves'
        measures_fpath = eval_dpath / 'curves/measures2.json'
        if measures_fpath.exists():
            data = json.loads(measures_fpath.read_text())
            pred_info = data['meta']['info']
        else:
            dvc_measures_fpath = measures_fpath.augment(tail='.dvc')
            if dvc_measures_fpath.exists():
                raise Exception('dvc pull')
            else:
                raise Exception('got nothing')

    param_types = parse_pred_params(pred_info, expt_dvc_dpath, path_hint)
    track_config = track_item['properties'].get('args', None)
    track_args = track_item['properties'].get('args', None)
    if track_config is not None:
        track_args = track_config
    track_config = relevant_track_config(track_args)
    param_types['track'] = track_config
    return param_types


def relevant_track_config(track_args):
    track_config = json.loads(track_args['track_kwargs'])
    track_config = {'trk_' + k: v for k, v in track_config.items()}
    return track_config


def parse_pred_params(pred_info, expt_dvc_dpath, path_hint=None):
    from watch.utils import util_time
    pred_item = find_pred_item(pred_info)

    # NOTE: the place where measure are stored has changed to be inside
    # the pred item.
    pred_measures = list(find_info_items(pred_info, 'measure', None))
    if len(pred_measures) > 0:
        # OLD CODE: Eventually delete
        assert len(pred_measures) <= 1
        meta = {'pred_start_time': None}
        if len(pred_measures):
            item = pred_measures[0]
            resources = parse_resource_item(item)
            predict_resources = item['properties']
            start_time = util_time.coerce_datetime(predict_resources.get('start_timestamp', None))
            meta['pred_start_time'] = start_time
        else:
            resources = {}
        fit_config = pred_item['properties']['fit_config']
    else:
        meta = {'pred_start_time': None}
        resources = {}
        # New code should have measures inside the pred item
        fit_config = pred_item['properties']['extra']['fit_config']
        meta['pred_start_time'] = pred_item['properties']['start_timestamp']
        meta['pred_end_time'] = pred_item['properties']['stop_timestamp']
        meta['start_timestamp'] = pred_item['properties']['start_timestamp']
        meta['stop_timestamp'] = pred_item['properties']['stop_timestamp']

        item = pred_item
        parse_resource_item(pred_item)
        resources = parse_resource_item(item)

    pred_config = pred_item['properties'].get('args', None)
    pred_args = pred_item['properties'].get('args', None)
    if pred_config is not None:
        pred_args = pred_config
    pred_config = relevant_pred_config(pred_args, expt_dvc_dpath)
    fit_config = relevant_fit_config(fit_config)

    param_types = {
        'fit': fit_config,
        'pred': pred_config,
        'resource': resources,
        'meta': meta,
    }
    return param_types


def parse_resource_item(item):
    from watch.utils import util_time
    ureg = global_ureg()
    pred_prop = item['properties']

    start_time = util_time.coerce_datetime(pred_prop.get('start_timestamp', None))
    end_time = util_time.coerce_datetime(pred_prop.get('end_timestamp', pred_prop.get('stop_timestamp', None)))
    iters_per_second = pred_prop.get('iters_per_second', None)
    total_hours = (end_time - start_time).total_seconds() / (60 * 60)

    try:
        vram = pred_prop['device_info']['allocated_vram']
        vram_gb = ureg.parse_expression(f'{vram} bytes').to('gigabytes').m
    except KeyError:
        vram_gb = None
    try:
        gpu_name = pred_prop['device_info']['device_name']
    except KeyError:
        gpu_name = None
    co2_kg = pred_prop['emissions']['co2_kg']

    if 'machine' in pred_prop:
        # New method
        cpu_name = pred_prop['machine']['cpu_brand']
        pred_prop['machine']['cpu_brand']
        co2_kg = pred_prop['emissions']['co2_kg']
        kwh = pred_prop['emissions']['total_kWH']
        disk_type = pred_prop['disk_info']['filesystem']
    else:
        kwh = None
        # Old method
        try:
            cpu_name = pred_prop['system_info']['cpu_info']['brand_raw']
        except KeyError:
            cpu_name = None
        try:
            disk_type = pred_prop['system_info']['disk_info']['filesystem']
        except KeyError:
            disk_type = None

    resources = {}
    resources['co2_kg'] = co2_kg
    resources['kwh'] = kwh
    resources['total_hours'] = total_hours
    resources['iters_per_second'] = iters_per_second
    resources['cpu_name'] = cpu_name
    resources['gpu_name'] = gpu_name
    resources['disk_type'] = disk_type
    resources['vram_gb'] = vram_gb

    import re
    cpu_name = re.sub('.*Gen Intel(R) Core(TM) ', '', cpu_name)
    resources['hardware'] = '{} {}'.format(cpu_name, gpu_name)
    return resources


def relevant_pred_config(pred_args, dvc_dpath):
    pred_config = {}
    pred_config['tta_fliprot'] = pred_args.get('tta_fliprot', 0)
    pred_config['tta_time'] = pred_args.get('tta_time', 0)
    pred_config['chip_overlap'] = pred_args['chip_overlap']
    pred_config['input_space_scale'] = pred_args.get('input_space_scale', None)
    pred_config['window_space_scale'] = pred_args.get('window_space_scale', None)
    pred_config['output_space_scale'] = pred_args.get('output_space_scale', None)
    pred_config['time_span'] = pred_args.get('time_span', None)
    pred_config['time_sampling'] = pred_args.get('time_sampling', None)
    pred_config['time_steps'] = pred_args.get('time_steps', None)
    pred_config['chip_dims'] = pred_args.get('chip_dims', None)
    pred_config['set_cover_algo'] = pred_args.get('set_cover_algo', None)
    pred_config['resample_invalid_frames'] = pred_args.get('resample_invalid_frames', None)
    pred_config['use_cloudmask'] = pred_args.get('use_cloudmask', None)
    package_fpath = pred_args['package_fpath']
    test_dataset = pred_args['test_dataset']
    if dvc_dpath is not None:
        package_fpath = resolve_cross_machine_path(package_fpath, dvc_dpath)
        test_dataset = resolve_cross_machine_path(test_dataset, dvc_dpath)
    # pred_config['model_fpath'] = package_fpath
    # pred_config['in_dataset'] = test_dataset
    pred_config['model_fpath'] = package_fpath
    pred_config['in_dataset_fpath'] = test_dataset

    pred_config['model_name'] = model_name = ub.Path(package_fpath).name
    pred_config['in_dataset_name'] = str(ub.Path(*test_dataset.parts[-2:]))

    # Hack to get the epoch/step/expt_name
    try:
        epoch = int(model_name.split('epoch=')[1].split('-')[0])
    except Exception:
        epoch = -1
    try:
        step = int(model_name.split('step=')[1].split('-')[0])
    except Exception:
        step = -1
    try:
        expt_name = model_name.split('_epoch=')[0]
    except Exception:
        expt_name = '?'
        # expt_name = predict_args[expt_name]

    pred_config = {'pred_' + k: v for k, v in pred_config.items()}
    pred_config['step'] = step
    pred_config['epoch'] = epoch
    pred_config['expt_name'] = expt_name
    return pred_config


def relevant_fit_config(fit_config):
    ignore_params = {
        'default_root_dir', 'enable_progress_bar'
        'prepare_data_per_node', 'enable_model_summary', 'checkpoint_callback',
        'detect_anomaly', 'gpus', 'terminate_on_nan', 'train_dataset',
        'workdir', 'config', 'num_workers', 'amp_backend',
        'enable_progress_bar', 'flush_logs_every_n_steps',
        'enable_checkpointing', 'prepare_data_per_node', 'amp_level',
        'vali_dataset', 'test_dataset', 'package_fpath', 'num_draw',
        'num_nodes', 'num_processes', 'num_sanity_val_steps',
        'overfit_batches', 'process_position',
        'reload_dataloaders_every_epoch', 'reload_dataloaders_every_n_epochs',
        'replace_sampler_ddp', 'sync_batchnorm', 'torch_sharing_strategy',
        'torch_start_method', 'val_check_interval', 'weights_summary',
        'auto_lr_find', 'auto_select_gpus', 'auto_scale_batch_size', 'benchmark',
        'check_val_every_n_epoch', 'draw_interval', 'eval_after_fit', 'fast_dev_run',
        'limit_predict_batches', 'limit_test_batches', 'limit_train_batches',
        'limit_val_batches', 'log_every_n_steps', 'logger',
        'move_metrics_to_cpu', 'multiple_trainloader_mode',
    }
    from scriptconfig import smartcast
    # hack, rectify different values of known parameters that mean the
    # same thing
    fit_config2 = ub.dict_diff(fit_config, ignore_params)
    for k, v in fit_config2.items():
        if k not in {'channels', 'init'}:
            v2 = smartcast.smartcast(v)
            if isinstance(v2, list):
                # Dont coerce into a list
                v2 = v
            fit_config2[k] = v2
        else:
            fit_config2[k] = v

    if 'init' in fit_config2:
        # hack to make init only use the filename
        fit_config2['init'] = fit_config2['init'].split('/')[-1]
    return fit_config2


def find_info_items(info, query_type, query_name=None):
    for item in info:
        if item['type'] == query_type:
            if query_name is None or item['properties']['name'] == query_name:
                yield item


def find_track_item(tracker_info):
    track_items = list(find_info_items(tracker_info, 'process', 'watch.cli.kwcoco_to_geojson'))
    assert len(track_items) == 1
    track_item = track_items[0]
    return track_item


def find_pred_item(pred_info):
    pred_items = list(find_info_items(pred_info, 'process', 'watch.tasks.fusion.predict'))
    assert len(pred_items) == 1
    pred_item = pred_items[0]
    return pred_item


def load_measure(measure_fpath):
    """
    Workers to load a single measure path. Has a hack to fix old configs.
    This can eventually be removed.
    """

    if measure_fpath.is_dir():
        cands = [
            measure_fpath / 'curves/measures2.json',
            measure_fpath / 'measures2.json',
        ]
        found = None
        for c in cands:
            if c.exists():
                found = c
                break
        if found is None:
            raise IOError(str(measure_fpath))
        measure_fpath = found

    HACK_FOR_IARPA = True
    if HACK_FOR_IARPA:
        import glob
        eval_dpath = measure_fpath.parent.parent
        iarpa_globs = eval_dpath / 'tracking/*/iarpa_eval/scores/merged/summary2.json'

        if 0:
            eval_parents = measure_fpath.parent.parent.parent
            cand = ub.Path(*['pred' if p == 'eval' else p for p in eval_parents.parts])
            iarpa_globs = cand / 'tracking/*/iarpa_eval/scores/merged/summary2.json'

        iarpa_summary_fpaths = list(glob.glob(str(iarpa_globs)))
        if ub.argflag('--force-iarpa'):
            if not len(iarpa_summary_fpaths):
                raise Exception('forcing-iarpa')
        # if len(iarpa_summary_fpaths) > 1:
        #     import xdev
        #     xdev.embed()
    else:
        iarpa_summary_fpaths = []

    with open(measure_fpath, 'r') as file:
        measure_info = json.load(file)

    measure_info['meta']['measure_fpath'] = measure_fpath

    if iarpa_summary_fpaths:
        iarpa_subresults = []
        for iarpa_fpath in iarpa_summary_fpaths:
            iarpa_fpath = ub.Path(iarpa_fpath)
            # HACK: need to persist the track params here
            with open(iarpa_fpath, 'r') as file:
                iarpa_merged = json.load(file)
                iarpa_subresults.append(iarpa_merged)
                # parent_info = iarpa_merged.get('parent_info', None)
                # sc_df = pd.read_json(io.StringIO(json.dumps(iarpa_merged['sc_df'])), orient='table')
                # sc_cm = pd.read_json(io.StringIO(json.dumps(iarpa_merged['sc_cm'])), orient='table')
        measure_info['meta']['iarpa_subresults'] = iarpa_subresults

    if True:
        # Hack to ensure fit config is properly serialized
        try:
            predict_meta = None
            for meta_item in measure_info['meta']['info']:
                if meta_item['type'] == 'process':
                    if meta_item['properties']['name'] == 'watch.tasks.fusion.predict':
                        predict_meta = meta_item
                        raise Found
        except Found:
            pass
        else:
            raise Exception('no prediction metadata')
        process_props = predict_meta['properties']
        predict_args = process_props['args']
        cand_remote = process_props['hostname']
        need_fit_config_hack = 'fit_config' not in process_props
        if need_fit_config_hack:
            # Hack, for models where I forgot to serialize the fit
            # configuration.
            print('Hacking in fit-config')
            package_fpath = predict_args['package_fpath']

            # TODO: use this hack instead resolve_cross_machine_path

            # hack, dont have enough into to properly remove the user directory
            hack_home = ub.expandpath(f'$HOME/remote/{cand_remote}')
            cand_remote_home = ub.Path(hack_home)
            tmp = ub.Path(package_fpath)
            possible_home_dirs = [
                ub.Path('/home/local/KHQ'),
                ub.Path('/home'),
            ]
            cand_suffix = None
            for possible_home in possible_home_dirs:
                possible_home_parts = possible_home.parts
                n = len(possible_home_parts)
                if tmp.parts[:n] == possible_home_parts:
                    cand_suffix = '/'.join(tmp.parts[n + 1:])
                    break
            if cand_suffix is None:
                raise Exception
            cand_remote_fpath = cand_remote_home / cand_suffix
            if cand_remote_fpath.exists():
                base_file = ub.zopen(cand_remote_fpath, ext='.pt')
                found = None
                for subfile in base_file.namelist():
                    if 'package_header/fit_config.yaml' in subfile:
                        found = subfile
                file = ub.zopen(cand_remote_fpath / found, ext='.pt')
                fit_config = yaml.safe_load(file)
                # TODO: this should have already existed
                process_props['fit_config'] = fit_config
                print('Backup measures: {}'.format(measure_fpath))
                shutil.copy(measure_fpath, ub.augpath(measure_fpath, suffix='.bak'))
                with open(measure_fpath, 'w') as file:
                    json.dump(measure_info, file)
            else:
                raise Exception
    return measure_info


def resolve_cross_machine_path(path, dvc_dpath=None):
    """
    HACK

    Attempt to determine what the local path to a file/directry would be
    if it exists on this machine. This assumes the path is something
    that was checked into DVC.

    Args:
        dvc_dpath : the preferred dvc dpath to associate the file with
            in case the older one points to multiple.
    """
    # pkg_dvc = SimpleDVC.find_root(package_fpath).resolve()
    # SimpleDVC.find_root(package_fpath).resolve()
    path = ub.Path(path)
    needs_resolve = not path.exists()
    if not needs_resolve:
        if dvc_dpath is not None:
            needs_resolve = not path.is_relative_to(dvc_dpath)

    if needs_resolve:
        expected_dnames = [
            'smart_watch_dvc',
            'smart_watch_dvc-hdd',
            'smart_watch_dvc-ssd',
        ]

        found_idx = None
        for dname in expected_dnames:
            try:
                idx = path.parts.index(dname)
            except ValueError:
                pass
            else:
                found_idx = idx
                break

        if found_idx is not None:
            # import watch
            # dvc_dpath = watch.find_smart_dvc_dpath()
            pname = ub.Path(*path.parts[idx + 1:])
            pname_dvc = pname.augment(tail='.dvc')
            cwd = ub.Path('.').absolute()
            candidates = []
            if dvc_dpath is not None:
                candidates.extend([
                    dvc_dpath / pname,
                    dvc_dpath / pname_dvc,
                ])
            candidates.extend([
                cwd / pname,
                cwd / pname_dvc,
            ])
            found = None
            try:
                for cand_path in candidates:
                    if cand_path.exists():
                        found = cand_path
                        if found.name.endswith('.dvc'):
                            found = found.augment(ext='')
                        raise Found
            except Found:
                pass
            if found:
                return found
    return path


@ub.memoize
def global_ureg():
    import pint
    ureg = pint.UnitRegistry()
    return ureg


def prepare_results(all_infos, coi_pattern, dvc_dpath=None):
    from kwcoco.coco_evaluator import CocoSingleResult
    from watch.utils import result_analysis
    from watch.utils import util_time
    ureg = global_ureg()

    class_rows = []
    mean_rows = []
    all_results = []
    results_list2 = []
    for info in ub.ProgIter(all_infos, desc='prepare results'):
        # Note: for now, the nocls part will refer to the saliency metrics and
        # the ovr part will be the class metrics. Should make per-head results
        # in the future.
        result = CocoSingleResult.from_json(info)
        all_results.append(result)

        class_aps = []
        class_aucs = []

        coi_aps = []
        coi_aucs = []

        meta = info['meta']

        predict_meta = None
        predict_props = None
        predict_args = None
        for meta_item in meta['info']:
            if meta_item['type'] == 'process':
                if meta_item['properties']['name'] == 'watch.tasks.fusion.predict':
                    if predict_meta is None:
                        predict_meta = meta_item
                        predict_props = predict_meta['properties']
                        predict_args = predict_props['args']
                    # raise Found
            resources = {}
            if meta_item['type'] == 'measure':
                predict_resources = meta_item['properties']
                start_time = util_time.coerce_datetime(predict_resources.get('start_timestamp', None))
                end_time = util_time.coerce_datetime(predict_resources.get('end_timestamp', predict_resources.get('stop_timestamp', None)))
                iters_per_second = predict_resources.get('iters_per_second', None)
                total_hours = (end_time - start_time).total_seconds() / (60 * 60)
                vram = predict_resources['device_info']['allocated_vram']
                vram_gb = ureg.parse_expression(f'{vram} bytes').to('gigabytes').m
                co2_kg = predict_resources['emissions']['co2_kg']
                resources['co2_kg'] = co2_kg
                resources['vram_gb'] = vram_gb
                resources['total_hours'] = total_hours
                resources['iters_per_second'] = iters_per_second

        if predict_args is None:
            raise Exception('no prediction metadata')

        pred_fpath = predict_args['pred_dataset']
        package_fpath = ub.Path(predict_args['package_fpath'])

        package_fpath = resolve_cross_machine_path(package_fpath, dvc_dpath)

        # _ = ub.Path(pred_fpath)
        # HACK
        # model_fpath = (_.parent.parent.parent / (_.parent.parent.name.split('pred_')[-1] + '.pt'))

        title = meta['title']

        if 'package_name' not in meta:
            if ' ' not in title:
                package_name = title
            else:
                raise AssertionError
        else:
            package_name = meta['package_name']

        if predict_args is not None:
            # Relevant prediction parameters also count as params
            pred_config = {}
            pred_config['pred_tta_fliprot'] = predict_args.get('tta_fliprot', 0)
            pred_config['pred_tta_time'] = predict_args.get('tta_time', 0)
            pred_config['pred_chip_overlap'] = predict_args['chip_overlap']
            pred_cfgstr = ub.hash_data(pred_config)[0:8]
            title = title + pred_cfgstr
            meta['prefix'] = package_name + pred_cfgstr

        # Hack to get the epoch/step/expt_name
        try:
            epoch = int(package_name.split('epoch=')[1].split('-')[0])
        except Exception:
            epoch = -1

        try:
            step = int(package_name.split('step=')[1].split('-')[0])
        except Exception:
            step = -1

        try:
            expt_name = package_name.split('epoch=')[0]
        except Exception:
            expt_name = predict_args[expt_name]

        salient_measures = info['nocls_measures']
        class_measures = info['ovr_measures']

        expt_class_rows = []
        coi_catnames = []
        for catname, bin_measure in class_measures.items():
            if coi_pattern.match(catname):
                coi_catnames.append(catname)
                coi_aps.append(bin_measure['ap'])
                coi_aucs.append(bin_measure['auc'])
            class_aps.append(bin_measure['ap'])
            class_aucs.append(bin_measure['auc'])
            class_row = {}
            class_row['AP'] = bin_measure['ap']
            class_row['AUC'] = bin_measure['auc']
            class_row['APUC'] = np.nanmean([bin_measure['ap'], bin_measure['auc']])
            class_row['catname'] = catname
            class_row['title'] = title
            class_row['package_name'] = package_name
            class_row['expt_name'] = expt_name
            class_row['epoch'] = epoch
            class_row['step'] = step
            class_row['pred_fpath'] = pred_fpath
            class_row['model_fpath'] = str(package_fpath)
            expt_class_rows.append(class_row)
        class_rows.extend(expt_class_rows)

        row = {}
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', 'Mean of empty slice')
            row['class_mAP'] = np.nanmean(class_aps) if len(class_aps) else np.nan
            row['class_mAUC'] = np.nanmean(class_aucs) if len(class_aucs) else np.nan
            row['class_mAPUC'] = np.nanmean([row['class_mAUC'], row['class_mAP']])

            row['coi_mAP'] = np.nanmean(coi_aps) if len(coi_aps) else np.nan
            row['coi_mAUC'] = np.nanmean(coi_aucs) if len(coi_aucs) else np.nan
            row['coi_mAPUC'] = np.nanmean([row['coi_mAUC'], row['coi_mAP']])

            row['salient_AP'] = salient_measures['ap']
            row['salient_AUC'] = salient_measures['auc']
            row['salient_APUC'] = np.nanmean([row['salient_AP'], row['salient_AUC']])

        row['catname'] = 'all'
        row['package_name'] = package_name
        row['title'] = title
        row['expt_name'] = expt_name
        row['epoch'] = epoch
        row['step'] = step
        row['pred_fpath'] = pred_fpath
        row['model_fpath'] = str(package_fpath)

        iarpa_subresults = meta.get('iarpa_subresults', [])

        iarpa_simplified = []
        for iarpa_merged in iarpa_subresults:
            try:
                # thresh = float(iarpa_fpath.parent.parent.parent.parent.name.split('_')[1].split('=')[1])
                # sc_cm = pd.read_json(io.StringIO(json.dumps(iarpa_merged['sc_cm'])), orient='table')
                # sc_df = pd.read_json(io.StringIO(json.dumps(iarpa_merged['sc_df'])), orient='table')
                best_bas_rows = pd.read_json(io.StringIO(json.dumps(iarpa_merged['best_bas_rows'])), orient='table')
                parent_info = iarpa_merged.get('parent_info', None)
                try:
                    bas_row = best_bas_rows.loc['merged'].reset_index().iloc[0].to_dict()
                except Exception:
                    bas_row = best_bas_rows[best_bas_rows['region_id'].isnull()].reset_index(drop=1).iloc[0].to_dict()

                # bas_f1 = bas_row['F1'].values.ravel()[0]
                track_kwargs = None
                # track_pred_info = None
                for pinfo in parent_info:
                    pinfo_type = pinfo.get('type', None)
                    if pinfo_type == 'process':
                        if pinfo['properties']['name'] == 'watch.cli.kwcoco_to_geojson':
                            _prop = pinfo['properties']
                            track_kwargs = json.loads(_prop['args']['track_kwargs'])
                            # track_pred_info = _prop['pred_info']  # these should agree
                # row['thresh'] = thresh
                # if track_pred_info != predict_props:
                #     pass

                BAS_metrics = {
                    'BAS_F1': bas_row['F1'],
                    'BAS_thresh': track_kwargs['thresh'],
                    'BAS_rho': bas_row['rho'],
                    'BAS_tau': bas_row['tau'],
                }
                iarpa_simplified.append(BAS_metrics)
                # row.update(BAS_metrics)
            except Exception:
                pass

        if iarpa_simplified:
            # TODO: need to expand out these rows for each pred parameter
            # setting
            BAS_metrics = max(iarpa_simplified, key=lambda x: x['BAS_F1'])
            row.update(BAS_metrics)
        else:
            BAS_metrics = None

        mean_rows.append(row)

        if predict_props is not None:
            cand_remote = predict_props['hostname']

            if 'fit_config' in predict_props:
                fit_config = predict_props['fit_config']
                # Add in defaults for new params
                fit_config.setdefault('normalize_perframe', False)
                result.meta['fit_config'] = fit_config
            else:
                raise Exception('Fit config was not serialized correctly')

            metrics = {
                'class_mAP': row['class_mAP'],
                'class_mAUC': row['class_mAUC'],
                'class_mAPUC': row['class_mAPUC'],
                'coi_mAP': row['coi_mAP'],
                'coi_mAUC': row['coi_mAUC'],
                'coi_mAPUC': row['coi_mAPUC'],
                'salient_AP': row['salient_AP'],
                'salient_AUC': row['salient_AUC'],
                'salient_APUC': row['salient_APUC'],
            }
            if BAS_metrics is not None:
                metrics['BAS_F1'] = BAS_metrics['BAS_F1']

            for class_row in expt_class_rows:
                metrics[class_row['catname'] + '_AP'] = class_row['AP']
                metrics[class_row['catname'] + '_AUC'] = class_row['AUC']

            # Add relevant train params here
            row['channels'] = fit_config['channels']
            row['time_steps'] = fit_config['time_steps']
            row['chip_size'] = fit_config['chip_size']

            row['pred_tta_time'] = pred_config['pred_tta_time']
            row['pred_tta_fliprot'] = pred_config['pred_tta_fliprot']
            row['pred_chip_overlap'] = pred_config['pred_chip_overlap']

            row['arch_name'] = fit_config['arch_name']
            row['normalize_perframe'] = fit_config.get('normalize_perframe', False)
            row['normalize_inputs'] = fit_config.get('normalize_inputs', False)
            row['train_remote'] = cand_remote

            # Hacks to normalize specific params
            def hack_smartcast(x):
                try:
                    return int(x)
                except Exception:
                    pass

            from scriptconfig import smartcast
            # hack, rectify different values of known parameters that mean the
            # same thing
            fit_config2 = {}
            for k, v in fit_config.items():
                if k not in {'channels', 'init'}:
                    v2 = smartcast.smartcast(v)
                    if isinstance(v2, list):
                        # Dont coerce into a list
                        v2 = v
                    fit_config2[k] = v2
                else:
                    fit_config2[k] = v

            if 'init' in fit_config2:
                # hack to make init only use the filename
                fit_config2['init'] = fit_config2['init'].split('/')[-1]

            fit_config = fit_config2
            # fit_config = ub.map_vals(smartcast.smartcast, fit_config)

            params = fit_config.copy()
            params.update(pred_config)

            result_meta = {
                'title': title,
                'epoch': epoch,
                'step': step,
                'expt_name': expt_name,
                'pred_fpath': pred_fpath,
                'model_fpath': package_fpath,
                'package_name': package_name,
                'coi_catnames': ','.join(sorted(coi_catnames)),
                'resources': resources,
            }

            result2 = result_analysis.Result(
                 name=result.meta['title'],
                 params=params,
                 metrics=metrics,
                 meta=result_meta
            )
            results_list2.append(result2)
    return class_rows, mean_rows, all_results, results_list2


def best_candidates(class_rows, mean_rows):
    # TOP CANDIDATE MODELS - FIND TOP K MODELS FOR EVERY METRIC
    # K = 10
    K = 4
    max_per_metric_per_expt = 2
    all_model_candidates = set()

    mean_metrics = [
        'coi_mAP', 'coi_mAUC', 'coi_mAPUC',
        'salient_AP', 'salient_AUC', 'salient_APUC'
    ]
    class_metrics = ['AP', 'AUC', 'APUC']

    subsets = {}
    if len(class_rows):
        class_df = pd.DataFrame(class_rows)
        class_candidate_indexes = []
        for class_metric in class_metrics:
            for catname, group in class_df.groupby('catname'):
                valid_indexes = []
                for expt_name, subgroup in group.groupby('expt_name'):
                    best_subgroup = subgroup.sort_values(class_metric, ascending=False).iloc[0:max_per_metric_per_expt]
                    valid_indexes.extend(best_subgroup.index.tolist())
                valid_indexes = sorted(set(valid_indexes))
                valid_group = group.loc[valid_indexes]
                top_group = valid_group.sort_values(class_metric, ascending=False).iloc[0:K]
                class_candidate_indexes.extend(top_group.index)
        top_class_indexes = sorted(set(class_candidate_indexes))
        class_subset = class_df.loc[top_class_indexes]
        subsets['class'] = class_subset = class_subset.sort_values('AP')
        all_model_candidates.update(set(class_subset['model_fpath'].tolist()))

    else:
        class_subset = []
        top_class_indexes = []

    if len(mean_rows):
        mean_df = pd.DataFrame(mean_rows)
        mean_candidate_indexes = []
        for metric in mean_metrics:
            valid_indexes = []
            for expt_name, subgroup in mean_df.groupby('expt_name'):
                best_subgroup = subgroup.sort_values(metric, ascending=False).iloc[0:max_per_metric_per_expt]
                valid_indexes.extend(best_subgroup.index.tolist())
            valid_indexes = sorted(set(valid_indexes))
            valid_group = mean_df.loc[valid_indexes]
            top_group = valid_group.sort_values(metric, ascending=False).iloc[0:K]
            mean_candidate_indexes.extend(top_group.index)
        top_mean_indexes = sorted(set(mean_candidate_indexes))
        mean_subset = mean_df.loc[top_mean_indexes]
        subsets['mean'] = mean_subset = mean_subset.sort_values('coi_mAPUC')
        all_model_candidates.update(set(mean_subset['model_fpath'].tolist()))

        sc_mean_subset = mean_subset[~mean_subset['coi_mAPUC'].isnull()].sort_values('coi_mAPUC')
        bas_mean_subset = mean_subset[~mean_subset['salient_APUC'].isnull()].sort_values('salient_APUC')
    else:
        mean_subset = []
        top_mean_indexes = []
        sc_mean_subset = []
        bas_mean_subset = []

    model_candidates = ub.ddict(list)
    pred_candidates = ub.ddict(list)

    if len(class_subset):
        print('Best Subset Table (Per-Class):')
        print(class_subset[class_metrics + ['catname', 'package_name']].to_string())
        model_candidates['sc'].append(class_subset['model_fpath'].values.tolist())
        pred_candidates['sc'].append(class_subset['pred_fpath'].values.tolist())

    if len(bas_mean_subset):
        print('Best Subset Table (Mean-BAS):')
        print(bas_mean_subset[mean_metrics + ['catname', 'package_name']].to_string())
        model_candidates['bas'].append(bas_mean_subset['model_fpath'].values.tolist())
        pred_candidates['bas'].append(bas_mean_subset['pred_fpath'].values.tolist())

    if len(sc_mean_subset):
        print('Best Subset Table (Mean-SC):')
        print(sc_mean_subset[mean_metrics + ['catname', 'package_name']].to_string())
        model_candidates['sc'].append(sc_mean_subset['model_fpath'].values.tolist())
        pred_candidates['sc'].append(sc_mean_subset['pred_fpath'].values.tolist())

    for n, s in subsets.items():
        print('n = {!r}'.format(n))
        print(shrink_notations(s, drop=1))

    sc_model_candidates = list(ub.unique(ub.flatten(model_candidates['sc'])))
    bas_model_candidates = list(ub.unique(ub.flatten(model_candidates['bas'])))

    sc_pred_candidates = list(ub.unique(ub.flatten(pred_candidates['sc'])))
    bas_pred_candidates = list(ub.unique(ub.flatten(pred_candidates['bas'])))

    print('sc_model_candidates = {}'.format(ub.repr2(sc_model_candidates, nl=1)))
    print('bas_model_candidates = {}'.format(ub.repr2(bas_model_candidates, nl=1)))

    print('sc_pred_candidates = {}'.format(ub.repr2(sc_pred_candidates, nl=1)))
    print('bas_pred_candidates = {}'.format(ub.repr2(bas_pred_candidates, nl=1)))

    all_model_candidates = sorted(all_model_candidates)

    # from watch.utils import util_path
    # for p in all_model_candidates:
    #     print(resolve_cross_machine_path(p, dvc_dpath))
    #     # print(util_path.resolve_directory_symlinks(resolve_cross_machine_path(p)))

    print('all_model_candidates = {}'.format(ub.repr2(all_model_candidates, nl=1)))

    if 0:
        # HACK
        # Check if models have predictions
        from watch.tasks.fusion.organize import suggest_paths
        for fpath in all_model_candidates:
            fpath = ub.Path(fpath)
            if fpath.exists():
                print('fpath = {!r}'.format(fpath))
                idx = fpath.parts.index('packages')
                prefix = ub.Path(*fpath.parts[:idx])
                prefix / 'pred'
                import glob
                suggestions = suggest_paths('*/*', package_fpath=fpath, as_json=0, pred_cfgstr='*', sidecar2=1)
                has_preds = list(glob.glob(suggestions['pred_dataset']))
                if has_preds:
                    print('has_preds = {!r}'.format(has_preds))
    return all_model_candidates


def shrink_channels(x):
    import kwcoco
    aliases = {
        'blue': 'B',
        'red': 'R',
        'green': 'G',
        'nir': 'N',
        'swir16': 'S',
        'swir22': 'H',
    }
    for idx, part in enumerate('forest|brush|bare_ground|built_up|cropland|wetland|water|snow_or_ice_field'.split('|')):
        aliases[part] =  'land.{}'.format(idx)
    stream_parts = []
    sensorchan = kwcoco.SensorChanSpec.coerce(x)
    # spec = kwcoco.ChannelSpec.coerce(x)
    for stream in sensorchan.streams():
        fused_parts = []
        for c in stream.chans.as_list():
            c = aliases.get(c, c)
            c = c.replace('matseg_', 'matseg.')
            fused_parts.append(c)
        fused = '|'.join(fused_parts)
        fused = fused.replace('B|G|R|N', 'BGRN')
        fused = fused.replace('B|G|R|N|S|H', 'BGRNSH')
        fused = fused.replace('R|G|B', 'RGB')
        fused = fused.replace('B|G|R', 'BGR')
        stream_parts.append(stream.sensor.spec + ':' + fused)
    new = ','.join(stream_parts)
    x = kwcoco.SensorChanSpec.coerce(new).concise().spec
    return x


def shrink_notations(df, drop=0):
    import kwcoco
    import re
    from watch.utils import util_regex
    b = util_regex.PythonRegexBuilder()
    pat0 = r'v\d+'
    pat1 = '^{pat}$'.format(pat=pat0)
    pat2 = b.lookbehind('_') + pat0 + b.optional(b.lookahead('_'))
    pat_text = b.oneof(*map(b.group, (pat1, pat2)))
    pat = re.compile(pat_text)

    shrunk = df.copy()

    if 0:
        shrunk['expt_name'] = (
            shrunk['expt_name'].apply(
                lambda x: pat.search(x).group()
            ))
    if 'channels' in shrunk:
        shrunk['channels'] = (
            shrunk['channels'].apply(
                lambda x: kwcoco.ChannelSpec.coerce(x.replace('matseg_', 'matseg.')).concise().spec
            ))
        shrunk['channels'] = (
            shrunk['channels'].apply(
                lambda x: x.replace('blue|green|red|nir|swir16|swir22', 'BGRNSH'))
        )
        shrunk['channels'] = (
            shrunk['channels'].apply(
                lambda x: x.replace('red|green|blue', 'RGB'))
        )
        shrunk['channels'] = (
            shrunk['channels'].apply(
                lambda x: x.replace('forest|brush|bare_ground|built_up|cropland|wetland|water|snow_or_ice_field', 'land:8'))
        )
        shrunk['channels'] = (
            shrunk['channels'].apply(
                lambda x: x.replace('brush|bare_ground|built_up', 'land:3'))
        )

    if drop:
        drop_cols = set(shrunk.columns) & {
            'title', 'normalize_perframe', 'normalize_inputs',
            'train_remote', 'step', 'arch_name', 'package_name',
            'pred_fpath', 'model_fpath',
        }
        shrunk = shrunk.drop(drop_cols, axis=1)
    return shrunk


def _oldhack():
    """
    if 0:
        # Hack to move over data into a comparable eval
        k1 = 'Drop1-Aligned-L1-2022-01_combo_DILM_nowv_vali.kwcoco'
        k2 = 'combo_DILM_nowv_vali.kwcoco'
        a = sorted(dset_groups[k1])
        b = sorted(dset_groups[k2])
        for x in b:
            y = ub.Path(str(x).replace(k2, k1))
            print(y in a)
            if not y.exists():
                p1 = x.parent.parent.parent
                p2 = y.parent.parent.parent
                ub.symlink(p1, p2, verbose=1)
    """


def hash_color(data):
    import distinctipy
    key_hash = ub.hash_data(data, hasher='blake3')
    key_tensor = np.frombuffer(memoryview(key_hash.encode()), dtype=np.int32)
    rng = kwarray.ensure_rng(rng=key_tensor.sum(), api='python')
    color = distinctipy.get_random_color(rng=rng)
    return color


def plot_summary_over_epochs(y, mean_df, dataset_title_part, fnum=1):
    import seaborn as sns
    import kwplot
    data = mean_df[~mean_df[y].isnull()]
    if len(data) == 0:
        return None
    kwplot.figure(fnum=fnum, doclf=True)
    ax = sns.lineplot(data=data, x='epoch', y=y, hue='expt_name', marker='o', style='channels')
    h, ell = ax.get_legend_handles_labels()
    ax.legend(h, ell, loc='lower right')
    ax.set_title(f'Pixelwise {y} metrics: {dataset_title_part}')  # todo: add train name
    # ax.set_title('Pixelwise mAP AC metrics: KR_R001 + KR_R002')
    fig = ax.figure
    return fig


def plot_individual_class_curves(all_results, dataset_title_part, catname, fnum, metric='ap'):
    from kwcoco.metrics import drawing
    import kwplot
    # max_num_curves = 16
    # max_num_curves = 32
    max_num_curves = 24
    # max_num_curves = 16
    max_per_expt = None
    # max_per_expt = 10
    max_per_expt = 3

    def lookup_metric(x):
        return x.ovr_measures[catname][metric]

    relevant_results = [r for r in all_results if catname in r.ovr_measures]
    # ub.group_items(relevant_results)

    if 1:
        # Take best per experiment
        groups = ub.group_items(relevant_results, key=lambda x: x.meta['fit_config']['name'])
        ordered_groups = []
        for name, group in groups.items():
            # if not ('v53' in name or 'v54' in name):
            #     continue
            ordered_group = sorted(group, key=lookup_metric)[::-1][:max_per_expt]
            ordered_groups.append(ordered_group)
        ordered_groups = sorted(ordered_groups, key=lambda g: lookup_metric(g[0]))[::-1]
        sorted_results = [x for x in ub.flatten(it.zip_longest(*ordered_groups)) if x is not None]
    else:
        sorted_results = sorted(relevant_results, key=lookup_metric)[::-1]

    results_to_plot = sorted_results[0:max_num_curves]
    results_to_plot = sorted(results_to_plot, key=lookup_metric)[::-1]
    # sorted_results = sorted(relevant_results, key=lookup_metric)[::-1]
    results_to_plot = sorted_results[0:max_num_curves]
    if len(results_to_plot) == 0:
        return None
    colors = kwplot.Color.distinct(len(results_to_plot))
    fig = kwplot.figure(fnum=fnum, doclf=True)
    for idx, result in enumerate(results_to_plot):
        color = colors[idx]
        color = [kwplot.Color(color).as01()]
        measure = result.ovr_measures[catname]
        if 'prefix' in result.meta:
            prefix = result.meta['prefix']
        elif 'package_name' in result.meta:
            prefix = result.meta['package_name']
        elif 'title' in result.meta:
            prefix = result.meta['title']
        else:
            prefix = '?label-unknown?'

        color = hash_color(prefix)

        kw = {'fnum': fnum}
        if metric == 'ap':
            drawing.draw_prcurve(measure, prefix=prefix, color=color, **kw)
        elif metric == 'auc':
            drawing.draw_roc(measure, prefix=prefix, color=color, **kw)
        else:
            raise KeyError
    fig.gca().set_title(f'Comparison of runs {metric}: {catname} -\n{dataset_title_part}')
    return fig


def plot_individual_salient_curves(all_results, dataset_title_part, fnum,
                                   metric='ap', rkey='nocls_measures',
                                   max_num_curves=24, max_per_expt=None):
    from kwcoco.metrics import drawing
    import kwplot
    # max_num_curves = 32
    # max_num_curves = 24
    # max_num_curves = 16
    # max_per_expt = None
    # max_per_expt = 10
    # max_per_expt = 3
    fig = kwplot.figure(fnum=fnum, doclf=True)
    relevant_results = [r for r in all_results if getattr(r, rkey) and getattr(r, rkey)['nsupport'] > 0]

    for result in relevant_results:
        if 'prefix' in result.meta:
            prefix = result.meta['prefix']
        elif 'package_name' in result.meta:
            prefix = result.meta['package_name']
        elif 'title' in result.meta:
            prefix = result.meta['title']
        else:
            prefix = '?label-unknown?'
        result.meta['prefix'] = prefix

    def lookup_metric(x):
        return getattr(x, rkey)[metric]

    if 1:
        # Take best per experiment
        groups = ub.group_items(relevant_results, key=lambda x: x.meta['fit_config']['name'])

        if 0:
            # HACK!!!!
            groups2 = {}
            for name in groups.keys():
                group = groups[name]
                if not ('v53' in name or 'v54' in name):
                    continue
                group2 = []
                for g in group:
                    flag1 = 'v53_epoch=15' in g.meta['prefix']
                    flag2 = 'v54_epoch=13' in g.meta['prefix']
                    if flag2 or flag1:
                        group2.append(g)
                group = group2
                if group:
                    groups2[name] = group
            groups = groups2
        ordered_groups = []
        for name, group in groups.items():
            ordered_group = sorted(group, key=lookup_metric)[::-1][:max_per_expt]
            ordered_groups.append(ordered_group)
        ordered_groups = sorted(ordered_groups, key=lambda g: lookup_metric(g[0]))[::-1]
        sorted_results = [x for x in ub.flatten(it.zip_longest(*ordered_groups)) if x is not None]
    else:
        sorted_results = sorted(relevant_results, key=lookup_metric)[::-1]

    results_to_plot = sorted_results[0:max_num_curves]
    results_to_plot = sorted(results_to_plot, key=lookup_metric)[::-1]

    if len(results_to_plot) == 0:
        return None

    colors = kwplot.Color.distinct(len(results_to_plot))
    for idx, result in enumerate(results_to_plot):
        color = colors[idx]
        color = [kwplot.Color(color).as01()]
        measure = getattr(result, rkey)
        prefix = result.meta['prefix']
        color = hash_color(prefix)
        kw = {'fnum': fnum}
        if metric == 'ap':
            drawing.draw_prcurve(measure, prefix=prefix, color=color, **kw)
        elif metric == 'auc':
            drawing.draw_roc(measure, prefix=prefix, color=color, **kw)
        else:
            raise KeyError
    fig.gca().set_title(f'Comparison of runs {metric}: {rkey} -\n{dataset_title_part}')
    return fig


def dump_spreadsheet(results_list2, out_dpath):
    spreadsheet_rows = [ub.dict_union(
        {'name': result.name},
        result.metrics,
        result.params,
        result.meta or {},
    )
        for result in results_list2]

    metrics_keys = set(ub.flatten(result.metrics.keys() for result in results_list2))

    ignore_spreadsheet = {
        'default_root_dir', 'enable_progress_bar'
        'prepare_data_per_node', 'enable_model_summary', 'checkpoint_callback',
        'detect_anomaly', 'gpus', 'terminate_on_nan',
        'workdir', 'config', 'num_workers', 'amp_backend',
        'enable_progress_bar', 'flush_logs_every_n_steps',
        'enable_checkpointing', 'prepare_data_per_node', 'amp_level',
        'package_fpath', 'num_draw',
        'track_grad_norm',
        'val_check_interval',
        'weights_summary',
        'process_position',
        'overfit_batches',
        'num_sanity_val_steps',
        'num_processes',
        'num_nodes',
        'move_metrics_to_cpu',
        'limit_val_batches',
        'limit_train_batches',
        'limit_predict_batches',
        'fast_dev_run',
        'eval_after_fit',
        'deterministic',
        'reload_dataloaders_every_epoch',
        'reload_dataloaders_every_n_epochs',
        'replace_sampler_ddp',
    }

    # https://pbpython.com/improve-pandas-excel-output.html
    # https://www.ojdo.de/wp/2019/10/pandas-to-excel-with-openpyxl/
    spreadsheet = pd.DataFrame(spreadsheet_rows)
    spreadsheet = spreadsheet.drop(set(spreadsheet.columns) & ignore_spreadsheet, axis=1)
    from openpyxl.formatting.rule import ColorScaleRule  # NOQA
    from openpyxl.styles import Alignment, Font, NamedStyle  # NOQA
    from openpyxl.utils import get_column_letter  # NOQA

    excel_fpath = out_dpath / 'experiment_results.xlsx'
    excel_fpath.delete()
    writer = pd.ExcelWriter(excel_fpath, engine='openpyxl', mode='w')
    with writer:
        spreadsheet.to_excel(writer, sheet_name='report', index=False)
        # workbook = writer.book
        ws = writer.sheets['report']

        ap_percentile_rule = ColorScaleRule(
            start_type='percentile',
            start_value=0,
            start_color='ffaaaa',  # red-ish
            mid_type='num',
            mid_value=0.3,
            mid_color='ffffff',  # white
            end_type='percentile',
            end_value=1,
            end_color='aaffaa')  # green-ish

        auc_percentile_rule = ColorScaleRule(
            start_type='percentile',
            start_value=0.4,
            start_color='ffaaaa',  # red-ish
            mid_type='num',
            mid_value=0.7,
            mid_color='ffffff',  # white
            end_type='percentile',
            end_value=1,
            end_color='aaffaa')  # green-ish

        metric_col_idxs = []

        for col_idx in range(1, ws.max_column):
            colname = spreadsheet.columns[col_idx - 1]
            col = get_column_letter(col_idx)
            if colname in metrics_keys:
                metric_col_idxs.append(col_idx)
            max_col_len = max(map(len, spreadsheet.iloc[:, (col_idx - 1)].to_string(index=False).split('\n')))
            # print('max_col_len = {!r}'.format(max_col_len))
            if max_col_len < 8:
                ws.column_dimensions[col].width = min(max(max_col_len, len(colname)), 26)
            else:
                ws.column_dimensions[col].width = 26

        # from matplotlib.colors import cmap
        # import matplotlib
        # cmap = matplotlib.cm.get_cmap('bwr')
        # cmap = matplotlib.cm.get_cmap('spectral')
        # metric_format = workbook.add_format({'num_format': '0.4f', 'bold': False})
        ws.column_dimensions['A'].width = 40
        for col_idx in metric_col_idxs:
            colname = spreadsheet.columns[col_idx - 1]
            col = get_column_letter(col_idx)
            value_cells = '{col}2:{col}{row}'.format(col=col, row=ws.max_row)
            ws.column_dimensions[col].width = 20
            if 'AUC' in colname:
                ws.conditional_formatting.add(value_cells, auc_percentile_rule)
            else:
                ws.conditional_formatting.add(value_cells, ap_percentile_rule)

            # Not working in google slides?
            # if 0:
            #     for row in ws[value_cells]:
            #         for cell in row:
            #             import kwimage
            #             try:
            #                 cell.fill.bgColor.rgb = #kwimage.Color(cmap(cell.value)).ashex()[1:7]
            #             except Exception:
            #                 pass
            #       # cell.fill.bgColor
            #       # cell.number_format = '0.0000'


def main(cmdline=False, **kwargs):
    """
    Ignore:
        from watch.tasks.fusion.aggregate_results import *  # NOQA
        import watch
        dvc_dpath = watch.find_smart_dvc_dpath()
        measure_globstr = 'models/fusion/SC-20201117/*/*/*/eval/curves/measures2.json'
        measure_globstr = 'models/fusion/SC-20201117/*_TA1*/*/*/eval/curves/measures2.json'
        cmdline = False
        kwargs = {}

        suffix = 'models/fusion/*/eval/*/*/*/*/eval/curves/measures2.json'
        kwargs['measure_globstr'] = dvc_dpath / suffix
        kwargs['out_dpath'] = dvc_dpath / '_agg_results2'

        if 0:
            remote = 'namek'
            remote_dpath = ub.Path(ub.shrinkuser(dvc_dpath, home=ub.expandpath(f'$HOME/remote/{remote}')))
            dvc_dpath = remote_dpath
    """
    from watch.utils import result_analysis
    from watch.utils import util_path
    from watch.utils import util_pattern

    config = AggregateResultsConfig(cmdline=cmdline, data=kwargs)
    print('config = {}'.format(ub.repr2(config.asdict(), nl=1)))

    measure_globstr = config['measure_globstr']
    out_dpath = ub.Path(config['out_dpath']).ensuredir()

    # TODO: high level results for a model should be serialized to DVC
    if measure_globstr is None:
        raise ValueError('Must specify a coercable glob pattern to locate the measures2.json files')
    else:
        measure_fpaths = util_path.coerce_patterned_paths(measure_globstr)

    if 0:
        pass

    coi_pattern = util_pattern.MultiPattern.coerce(
        config['classes_of_interest'], hint='glob')

    measure_fpaths = [ub.Path(p) for p in measure_fpaths]

    # FIXME
    # HACK: relies on directory structure
    HACK = 'EVAL2'
    HACK = 'EVAL3'
    if HACK == 'EVAL2':
        dset_groups = ub.group_items(
            measure_fpaths,
            lambda x: x.parent.parent.parent.name
        )
    elif HACK == 'EVAL3':
        dset_groups = ub.group_items(
            measure_fpaths,
            lambda x: x.parent.parent.parent.parent.name
        )
    else:
        raise NotImplementedError('fixme')
    # measure_fpaths[0].parent.parent.parent.parent

    # dset_glob = config['dset_glob']
    # dset_glob = ''
    # measure_fpaths

    print('dset_groups = {}'.format(ub.repr2(dset_groups, nl=2)))

    predict_group_freq = ub.map_vals(len, dset_groups)
    print('These are the different datasets prediction was run on.')
    print('TODO: need to choose exactly 1 or a compatible set of them')
    print('predict_group_freq = {}'.format(ub.repr2(predict_group_freq, nl=1)))

    dset_group_key = util_pattern.MultiPattern.coerce(
        config['dset_group_key'], hint='glob')
    dataset_keys = [k for k in dset_groups.keys()
                    if dset_group_key.match(k)]
    measure_fpaths = list(ub.flatten([dset_groups[k] for k in dataset_keys]))

    print(len(measure_fpaths))
    # dataset_to_evals = ub.group_items(eval_dpaths, lambda x: x.parent.name)

    load_workers = config['io_workers']
    jobs = ub.JobPool('thread', max_workers=load_workers)
    all_infos = []
    for measure_fpath in ub.ProgIter(measure_fpaths):
        job = jobs.submit(load_measure, measure_fpath)
        job.measure_fpath = measure_fpath

    # job = next(iter(jobs.as_completed(desc='collect jobs')))
    # all_infos = [job.result()]
    failed_jobs = []
    for job in jobs.as_completed(desc='collect jobs'):
        try:
            measure_info = job.result()
            all_infos.append(measure_info)
        except Exception as ex:
            failed_jobs.append(job.measure_fpath)
            print('Failed job.measure_fpath = {!r} because {!r}'.format(job.measure_fpath, ex))
            pass

    if 0:
        to_unlink = []
        for fpath in failed_jobs:
            if 'Drop2' in fpath.parent.parent.parent.name:
                dvc_fpath = ub.Path(str(fpath) + '.dvc')
                if dvc_fpath.exists():
                    to_unlink.append(fpath)
                    to_unlink.append(dvc_fpath)
                    print('fpath = {!r}'.format(fpath))
            print(len(list(fpath.parent.parent.parent.glob('*'))))
        for p in to_unlink:
            p.unlink()

    print(f'Failed Jobs {len(failed_jobs)=}/{len(jobs)}')

    from watch.utils.simple_dvc import SimpleDVC
    dvc = SimpleDVC.coerce(out_dpath)
    dvc_dpath = dvc.dvc_root

    class_rows, mean_rows, all_results, results_list2 = prepare_results(all_infos, coi_pattern, dvc_dpath=dvc_dpath)

    if 1:
        dump_spreadsheet(results_list2, out_dpath)

    if 0:
        best_candidates(class_rows, mean_rows)

    ignore_params = {
        'default_root_dir', 'name', 'enable_progress_bar'
        'prepare_data_per_node', 'enable_model_summary', 'checkpoint_callback',
        'detect_anomaly', 'gpus', 'terminate_on_nan', 'train_dataset',
        'workdir', 'config', 'num_workers', 'amp_backend',
        'enable_progress_bar', 'flush_logs_every_n_steps',
        'enable_checkpointing', 'prepare_data_per_node', 'amp_level',
        'vali_dataset', 'test_dataset', 'package_fpath',
        'num_draw',
    }
    ignore_metrics = {
        'positive_AUC',
        'positive_AP',
        # 'nocls_auc',
        # 'nocls_ap',
        # 'map',
        # 'mauc',
    }

    FILTER_OUTLIERS = 0
    # r = results_list2[0]
    if FILTER_OUTLIERS:
        # Filter outliers by only taking the top results for each experiment
        expt_name_to_results2 = ub.group_items(results_list2, lambda r: r.meta['expt_name'])
        results3 = []
        for expt_name, sub_results in expt_name_to_results2.items():
            top = sorted(sub_results, key=lambda r: r.metrics['salient_AP'], reverse=True)[0:1]
            results3.extend(top)
        results_input = results3
    else:
        results_input = results_list2
    abalation_orders = {1}
    analysis = result_analysis.ResultAnalysis(
        results_input, ignore_params=ignore_params,
        # metrics=['coi_mAPUC', 'coi_APUC'],
        # metrics=['salient_AP'],
        metrics=['coi_mAP', 'salient_AP'],
        metric_objectives={
            'salient_AP': 'max',
            'coi_mAP': 'max',
        },
        ignore_metrics=ignore_metrics,
        abalation_orders=abalation_orders
    )
    try:
        analysis.run()
    except TypeError:
        raise
    except Exception as ex:
        print('AnalysisError: ex = {!r}'.format(ex))
        print('Warning: Statistical analysis failed. Probably needs more data.')
    else:
        print('analysis.varied = {}'.format(ub.repr2(analysis.varied, nl=2)))
        if len(analysis.stats_table):
            analysis.stats_table = analysis.stats_table.sort_values('anova_rank_p')
            print(analysis.stats_table)

    class_df = pd.DataFrame(class_rows)
    mean_df = pd.DataFrame(mean_rows)
    class_df = class_df.drop(set(class_df.columns) & {'title', 'pred_fpath', 'package_name'}, axis=1)
    mean_df = mean_df.drop(set(mean_df.columns) & {'title', 'pred_fpath', 'package_name'}, axis=1)

    # metric_correlation
    metrics_of_interest = [
        'BAS_F1',
        'salient_AUC',
        'salient_AP',
        'coi_mAUC',
        'coi_mAP',
        'class_mAP',
        'class_mAUC',
    ]
    metrics_avail = mean_df.columns.intersection(metrics_of_interest)
    metrics_unavail = ub.oset(metrics_of_interest) - ub.oset(mean_df.columns)
    print('metrics_avail = {!r}'.format(metrics_avail.tolist()))
    print('metrics_unavail = {!r}'.format(list(metrics_unavail)))

    metric_corr = mean_df[metrics_avail].corr()
    print('Metric correleation')
    print(metric_corr.to_string())

    if 0:
        import kwplot
        sns = kwplot.autosns()
        sns.scatterplot(data=mean_df, x='salient_AP', y='BAS_F1')
        sns.scatterplot(data=mean_df, x='salient_AP', y='salient_AUC')

    mean_desc = mean_df.describe().T
    _nnull = mean_df.isnull().sum()
    mean_desc = mean_desc.assign(null=_nnull)
    print(mean_desc.to_string())

    class_df = shrink_notations(class_df, drop=1)
    mean_df = shrink_notations(mean_df, drop=1)

    def sort_via(label, df):
        if label in df.columns:
            _df_via_metric = df.sort_values(label)
            _df_via_metric = _df_via_metric[~_df_via_metric[label].isnull()]
            if len(_df_via_metric):
                print(f'\nSort by {label}')
                print(_df_via_metric.to_string())

    sort_via('coi_mAPUC', mean_df)
    sort_via('salient_APUC', mean_df)
    sort_via('BAS_F1', mean_df)
    sort_via('AP', class_df)

    # if 'coi_mAPUC' in mean_df.columns:
    #     print('\nSort by coi_mAPUC')
    #     _mean_by_metric = mean_df.sort_values('coi_mAPUC')
    #     # print(_mean_by_metric)
    #     print(_mean_by_metric.to_string())

    # if 'salient_APUC' in mean_df.columns:
    #     print('\nSort by salient_APUC')
    #     _salient_by_metric = mean_df.sort_values('salient_APUC')
    #     # print(_salient_by_metric)
    #     print(_salient_by_metric.to_string())

    # if 'BAS_F1' in mean_df.columns:
    #     print('\nSort by BAS_F1')
    #     _salient_by_metric = mean_df.sort_values('BAS_F1')
    #     # print(_salient_by_metric)
    #     print(_salient_by_metric.to_string())

    # if 'AP' in class_df.columns:
    #     print('\nClass: Sort by AP')
    #     _class_by_metric = class_df[~class_df['AP'].isnull()].sort_values('AP')
    #     # print(_class_by_metric)
    #     print(_class_by_metric.to_string())

    if 0:
        if 'AUC' in class_df.columns:
            print('\nClass: Sort by AUC')
            print(class_df[~class_df['AUC'].isnull()].sort_values('AUC').to_string())

    # mean_df['title'].apply(lambda x: int(x.split('epoch=')[1].split('-')[0]))
    def group_by_best(mean_df, metric_key, shrink=False):
        bests = []
        if len(mean_df):
            for t, subdf in mean_df.groupby('expt_name'):
                idx = subdf[[metric_key]].idxmax()
                import math
                if not math.isnan(idx.item()):
                    best = subdf.loc[idx]
                    bests.append(best)
        best_per_expt = pd.concat(bests)
        if shrink:
            best_per_expt = shrink_notations(best_per_expt, drop=1)
            best_per_expt = best_per_expt.rename({
                'time_steps': 'time',
                'chip_size': 'space'}, axis=1)
        return best_per_expt

    print('\nBest Class Models')
    try:
        best_per_expt = group_by_best(mean_df, 'coi_mAP', shrink=True)
        best_per_expt = best_per_expt[~best_per_expt['coi_mAP'].isnull()]
        print(best_per_expt.sort_values('coi_mAP').to_string())

        if 0:
            import dataframe_image as dfi
            dfi.export(
                best_per_expt,
                "./tmp.png",
                table_conversion="chrome",
                fontsize=28,
                max_rows=-1,
            )
            import kwplot
            fig, _ = kwplot.imshow('./tmp.png', fnum=10)
            fig.tight_layout()
    except ValueError:
        pass

    # salient_metric = 'salient_APUC'
    # class_metric = 'coi_mAPUC'
    salient_metric = 'salient_AP'
    class_metric = 'coi_mAP'

    try:
        print('\nBest Salient Models')
        best_per_expt = group_by_best(mean_df, salient_metric, shrink=True)
        print(best_per_expt.sort_values(salient_metric).to_string())
    except (ValueError, KeyError):
        pass

    _ = best_candidates(class_rows, mean_rows)

    if config['embed']:
        import xdev
        xdev.embed()

    DRAW = 1
    if DRAW:
        import kwplot
        sns = kwplot.autosns()  # NOQA
        plt = kwplot.autoplt()  # NOQA

        dataset_title_part = "-".join(dataset_keys)
        figsize = 'auto'
        verbose = 1
        if figsize == 'auto':
            figsize = (9, 7)

        if len(mean_df):
            y = class_metric
            fig1 = plot_summary_over_epochs(y, mean_df, dataset_title_part, fnum=1)
            if fig1 is not None:
                _writefig(fig1, out_dpath, 'epoch_summary_class.png', figsize, verbose, tight=True)

            y = salient_metric
            fig2 = plot_summary_over_epochs(y, mean_df, dataset_title_part, fnum=2)
            if fig2 is not None:
                _writefig(fig2, out_dpath, 'epoch_summary_salient.png', figsize, verbose, tight=True)

            # kwplot.figure(fnum=2, doclf=True)
            # ax = sns.lineplot(data=mean_df, x='epoch', y='class_mAUC', hue='expt_name', marker='o', style='channels')
            # ax.set_title('Pixelwise mAUC AC metrics: KR_R001 + KR_R002')

        if len(all_results):
            fnum = 3

            for result in all_results:
                coi_measures = []
                for c, m in result.ovr_measures.items():
                    if coi_pattern.match(c):
                        coi_measures.append(m)
                # Is this actually macro or micro? I forget
                import kwcoco
                if coi_measures:
                    ovr_coi_macro = kwcoco.metrics.Measures.combine(coi_measures)
                    ovr_coi_macro = ovr_coi_macro.reconstruct()
                    ovr_coi_macro.proxy['node'] = 'OVR-COI-Macro'
                    result.ovr_coi_macro = ovr_coi_macro
                else:
                    result.ovr_coi_macro = None

            catname = 'Active Construction'
            fig3 = plot_individual_class_curves(all_results, dataset_title_part, catname, fnum, 'ap')
            if fig3 is not None:
                _writefig(fig3, out_dpath, f'{catname}_ap_curve.png', figsize, verbose, tight=True)

            fnum = 4
            catname = 'Site Preparation'
            fig4 = plot_individual_class_curves(all_results, dataset_title_part, catname, fnum, 'ap')
            if fig4 is not None:
                _writefig(fig4, out_dpath, f'{catname}_ap_curve.png', figsize, verbose, tight=True)

        if len(all_results):
            fnum = 5
            fig5 = plot_individual_salient_curves(all_results, dataset_title_part, fnum, metric='ap', rkey='nocls_measures')
            if fig5 is not None:
                _writefig(fig5, out_dpath, 'salient_ap_curve.png', figsize, verbose, tight=True)
            # print(best_per_expt.sort_values('mAP').to_string())

            fnum = 6
            fig5 = plot_individual_salient_curves(all_results, dataset_title_part, fnum, metric='ap', rkey='ovr_coi_macro', max_per_expt=config['max_per_expt_salient_curves'])
            if fig5 is not None:
                _writefig(fig5, out_dpath, 'ovr_coi_macro_ap_curve.png', figsize, verbose, tight=True)

        if 0:
            # Make robust
            resource_rows = []
            # Resource scatter plots
            for r in results_list2:
                row = r.meta['resources'].copy()
                row.update(r.params)
                row['salient_AP'] = r.metrics['salient_AP']
                resource_rows.append(row)
            if resource_rows:
                resource_df = pd.DataFrame(resource_rows)

                fig = kwplot.figure(fnum=6)
                fig.clf()
                ax = fig.gca()
                x = 'total_hours'
                # x = 'co2_kg'
                # x = 'vram_gb'
                # x = 'iters_per_second'
                y = 'salient_AP'
                text_column = ['pred_tta_fliprot', 'pred_tta_time', 'pred_chip_overlap']
                ax = sns.scatterplot(data=resource_df, x=x, y=y, ax=ax)
                # Add text besides each point
                data = resource_df
                for line in range(0, data.shape[0]):
                    xcoord = data[x][line]
                    ycoord = data[y][line]
                    subcol = data[text_column].iloc[line]
                    text = ub.repr2(subcol.to_dict(), compact=1).replace('pred_', '')
                    ax.text(xcoord, ycoord,
                            text, horizontalalignment='left',
                            size='medium', color='black', weight='semibold')
                    ax.set_title(f'Time/Accuracy Tradeoffs: {dataset_title_part}')  # todo: add train name

        # # if 1:
        #     # fig3.set_size_inches(np.array([6.4, 4.8]) * 2.0)
        #     # fig3.tight_layout()
        #     # fig4.set_size_inches(np.array([6.4, 4.8]) * 2.0)
        #     # fig4.tight_layout()
        #     # fig5.set_size_inches(np.array([5.4, 2.8]) * 2.0)
        #     # fig5.tight_layout()

        if config['show']:
            plt.show()


if __name__ == '__main__':
    """
    CommandLine:
        # DVC_DPATH=$(smartwatch_dvc)
        DVC_DPATH=$HOME/data/dvc-repos/smart_watch_dvc
        MEASURE_GLOBSTR=$DVC_DPATH/models/fusion/SC-20201117/*_TA1*/*/*/eval/curves/measures2.json
        python -m watch.tasks.fusion.aggregate_results \
            --measure_globstr="$MEASURE_GLOBSTR" \
            --out_dpath="$DVC_DPATH/agg_results"
            --dset_group_key="Drop2-Aligned-TA1-2022-01_*.kwcoco"


        DVC_DPATH=$HOME/data/dvc-repos/smart_watch_dvc
        cd $DVC_DPATH/models/fusion/SC-20201117
        python -m watch.tasks.fusion.aggregate_results \
            --measure_globstr="*/*/*/eval/curves/measures2.json" \
            --out_dpath="$DVC_DPATH/agg_results" \
            --dset_group_key="*" --show=True
    """
    main(cmdline=True)
